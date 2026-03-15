"""
CFTC Weekly Swaps Report Parser

Parses weekly swap reports from CFTC containing interest rate, credit, and FX
swap notional outstanding, cleared/uncleared splits, and counterparty breakdowns.

Each file contains 52 sheets with weekly snapshots. Sheet 1 is the overview
with all asset classes. Values are in millions USD.

Data: ~600 weekly files (2013-2026).
"""

import os
import re
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime

ASSET_CLASS_PATTERNS = [
    (re.compile(r'total\s+interest\s+rate', re.IGNORECASE), 'ir'),
    (re.compile(r'total\s+cross.?currency', re.IGNORECASE), 'cross_currency'),
    (re.compile(r'total\s+credit', re.IGNORECASE), 'credit'),
    (re.compile(r'total\s+fx', re.IGNORECASE), 'fx'),
    (re.compile(r'total\s+equit', re.IGNORECASE), 'equity'),
    (re.compile(r'total\s+commodit', re.IGNORECASE), 'commodity'),
]

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'raw', 'swaps')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'processed')


def _extract_date_from_filename(filename):
    """Extract date from CFTC_Swaps_Report_MM_DD_YYYY.xlsx"""
    name = filename.replace('.xlsx', '').replace('cftc_swaps_report_', '').replace(
        'CFTC_Swaps_Report_', '')
    parts = name.split('_')
    try:
        if len(parts) >= 3:
            month, day, year = int(parts[0]), int(parts[1]), int(parts[2])
            if year < 100:
                year += 2000
            return datetime(year, month, day)
    except (ValueError, IndexError):
        pass
    return None


def parse_overview_sheet(filepath):
    """Parse Sheet 1 (overview) — IR, Credit, FX notional with weekly columns.

    Returns a DataFrame with columns: date, metric, value (millions USD).
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  WARNING: Could not open {os.path.basename(filepath)}: {e}")
        return pd.DataFrame()

    if '1' not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb['1']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    # Row 0: header with dates
    header = rows[0]

    # Infer year from filename for files with abbreviated date headers
    file_date = _extract_date_from_filename(os.path.basename(filepath))

    # Extract dates from header
    # Some Excel files return dates as serial numbers (floats) instead of datetime
    _excel_epoch = datetime(1899, 12, 30)
    dates = []
    for val in header[1:]:
        if val is None:
            dates.append(None)
        elif isinstance(val, datetime):
            dates.append(val)
        elif isinstance(val, (int, float)) and 30000 < val < 60000:
            # Excel serial date number
            dates.append(_excel_epoch + pd.Timedelta(days=int(val)))
        else:
            try:
                dates.append(pd.to_datetime(val))
            except Exception:
                # Try abbreviated dates like "Dec 7" or "January 3" (early 2013-2014)
                if file_date and isinstance(val, str):
                    try:
                        # Parse month+day, infer year from filename
                        parsed = pd.to_datetime(val, format='mixed')
                        dates.append(parsed.replace(year=file_date.year))
                    except Exception:
                        for fmt in ('%B %d', '%b %d'):
                            try:
                                parsed = datetime.strptime(val.strip(), fmt)
                                # Use filename year; adjust if month > filename month
                                # (e.g. file is Jan 2013 but column is "Dec" = Dec 2012)
                                year = file_date.year
                                if parsed.month > file_date.month + 1:
                                    year -= 1
                                dates.append(parsed.replace(year=year))
                                break
                            except ValueError:
                                continue
                        else:
                            dates.append(None)
                else:
                    dates.append(None)

    # Dynamic label-based row mapping — handles all CFTC layout eras
    records = []
    current_class = None

    for row in rows[1:]:
        label = str(row[0] or '').strip()
        label_clean = label.lower().replace('*', '').strip()

        # Skip empty rows, grand total, and footnote text
        if not label_clean or label_clean == 'total' or len(label_clean) > 60:
            current_class = None
            continue

        # Check if this is an asset class header row (e.g. "Total Interest Rate")
        matched = False
        for pattern, prefix in ASSET_CLASS_PATTERNS:
            if pattern.search(label_clean):
                current_class = prefix
                metric = f'{prefix}_total'
                matched = True
                break

        if not matched and current_class:
            if 'uncleared' in label_clean:
                metric = f'{current_class}_uncleared'
            elif 'cleared' in label_clean:
                metric = f'{current_class}_cleared'
            else:
                continue  # Unknown sub-row, skip
        elif not matched:
            continue

        # Extract values for all date columns
        for col_idx, date in enumerate(dates):
            if date is None:
                continue
            val = row[col_idx + 1] if col_idx + 1 < len(row) else None
            if val is not None:
                records.append({
                    'date': date,
                    'metric': metric,
                    'value_millions': pd.to_numeric(val, errors='coerce'),
                })

    return pd.DataFrame(records)


def parse_all_swaps(data_dir=None, output_dir=None):
    """Parse all weekly swap reports and produce processed CSVs."""
    if data_dir is None:
        data_dir = DATA_DIR
    if output_dir is None:
        output_dir = OUTPUT_DIR

    os.makedirs(output_dir, exist_ok=True)

    files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx')])
    print(f"Parsing {len(files)} CFTC weekly swap reports...")

    all_records = []
    failed_files = []

    for i, f in enumerate(files):
        filepath = os.path.join(data_dir, f)
        df = parse_overview_sheet(filepath)

        if df.empty:
            failed_files.append(f)
            continue

        all_records.append(df)

        if (i + 1) % 50 == 0:
            print(f"  [{i+1}/{len(files)}] processed...")

    if not all_records:
        print("No swap data parsed!")
        return

    # Combine all records and deduplicate (files contain overlapping weekly dates)
    combined = pd.concat(all_records, ignore_index=True)
    combined = combined.drop_duplicates(subset=['date', 'metric'], keep='last')
    combined = combined.sort_values(['date', 'metric'])

    # Convert to billions for consistency
    combined['value_billions'] = combined['value_millions'] / 1000

    # --- Pivot to wide format for easier analysis ---
    wide = combined.pivot_table(index='date', columns='metric',
                                values='value_billions', aggfunc='first')
    wide = wide.reset_index()
    wide.columns.name = None

    # Compute derived metrics
    if 'ir_total' in wide.columns and 'ir_cleared' in wide.columns:
        wide['ir_cleared_pct'] = wide['ir_cleared'] / wide['ir_total']
    if 'credit_total' in wide.columns and 'credit_cleared' in wide.columns:
        wide['credit_cleared_pct'] = wide['credit_cleared'] / wide['credit_total']
    if 'fx_total' in wide.columns and 'fx_cleared' in wide.columns:
        wide['fx_cleared_pct'] = wide['fx_cleared'] / wide['fx_total']
    if 'cross_currency_total' in wide.columns and 'cross_currency_cleared' in wide.columns:
        wide['cross_currency_cleared_pct'] = wide['cross_currency_cleared'] / wide['cross_currency_total']
    if 'equity_total' in wide.columns and 'equity_cleared' in wide.columns:
        wide['equity_cleared_pct'] = wide['equity_cleared'] / wide['equity_total']

    # --- Save weekly time series ---
    wide.to_csv(os.path.join(output_dir, 'swaps_weekly.csv'), index=False)
    print(f"  Saved swaps_weekly.csv ({len(wide)} rows)")

    # --- Save long format ---
    combined.to_csv(os.path.join(output_dir, 'swaps_weekly_long.csv'), index=False)
    print(f"  Saved swaps_weekly_long.csv ({len(combined)} rows)")

    # --- Quarterly aggregation ---
    wide['quarter'] = pd.to_datetime(wide['date']).dt.to_period('Q').astype(str)
    q_agg = {}
    for col in wide.columns:
        if col in ['date', 'quarter']:
            continue
        q_agg[col] = 'last'  # Use quarter-end value
    q_agg[f'ir_total_mean'] = ('ir_total', 'mean') if 'ir_total' in wide.columns else None

    quarterly = wide.groupby('quarter').agg(
        weeks=('date', 'count'),
        **{col: (col, 'last') for col in wide.columns if col not in ['date', 'quarter']}
    ).reset_index()
    quarterly.to_csv(os.path.join(output_dir, 'swaps_quarterly.csv'), index=False)
    print(f"  Saved swaps_quarterly.csv ({len(quarterly)} rows)")

    # --- Summary ---
    print(f"\nDone! {len(files)} files parsed, {len(failed_files)} failed.")
    if failed_files:
        print("  Failed files:")
        for ff in failed_files[:20]:
            print(f"    - {ff}")
        if len(failed_files) > 20:
            print(f"    ... and {len(failed_files) - 20} more")
    latest = wide.iloc[-1]
    print(f"  Latest date: {latest['date']}")
    if 'ir_total' in wide.columns:
        print(f"  IR notional: ${latest['ir_total']:,.0f}B (cleared: {latest.get('ir_cleared_pct', 0):.1%})")
    if 'credit_total' in wide.columns:
        print(f"  Credit notional: ${latest['credit_total']:,.0f}B (cleared: {latest.get('credit_cleared_pct', 0):.1%})")
    if 'fx_total' in wide.columns:
        print(f"  FX notional: ${latest['fx_total']:,.0f}B (cleared: {latest.get('fx_cleared_pct', 0):.1%})")


if __name__ == '__main__':
    parse_all_swaps()
